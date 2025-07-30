import pdfplumber
import pandas as pd
import re
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

path_pdf = r"C:\Users\User\Downloads\EXTRATO DE DÉBITOS COMPLETO.pdf"
dados = []

def extrair_responsavel(bloco_texto):
    texto = ' '.join(bloco_texto).replace('\n', ' ')
    nome_match = re.search(r"^(.*?) - FONE:", texto)
    cpf_match = re.search(r"CPF:\s*([\d]+)", texto)
    email_match = re.search(r"E-MAIL:\s*([^\s]+)", texto)
    return {
        "nome": nome_match.group(1).title().strip() if nome_match else "",
        "cpf": cpf_match.group(1) if cpf_match else "",
        "email": email_match.group(1) if email_match else ""
    }

with pdfplumber.open(path_pdf) as pdf:
    responsavel = {"nome": "", "cpf": "", "email": ""}
    for page in pdf.pages:
        linhas = page.extract_text().split('\n')
        i = 0
        while i < len(linhas):
            linha = linhas[i].strip()

            if linha.startswith("DADOS DO RESPONSÁVEL FINANCEIRO"):
                i += 1
                bloco = []
                while i < len(linhas):
                    l = linhas[i].strip()
                    if not l or l.startswith(("DADOS DO RESPONSÁVEL FINANCEIRO","DÉBITOS EM ABERTO")):
                        break
                    bloco.append(l)
                    i += 1
                novo = extrair_responsavel(bloco)
                if novo["nome"]:
                    responsavel = novo
                continue

            if linha.startswith("DÉBITOS EM ABERTO"):
                i += 1
                bloco = []
                while i < len(linhas):
                    l = linhas[i].strip()
                    if not l or l.startswith(("DADOS DO RESPONSÁVEL FINANCEIRO","VALOR TOTAL")):
                        break
                    bloco.append(l)
                    i += 1

                texto_debitos = "\n".join(bloco)
                padrao = re.compile(
                    r"(\d{7}) - ([^\n]+)\n"            # código + nome parcial
                    r"(\d+º Parcela) (\d{2}/\d{2}/\d{4}) .*? R\$ ([\d.,]+) R\$ ([\d.,]+)\n"
                    r"([^\n]+)?",                      # linha complementar opcional
                    re.MULTILINE
                )

                for m in padrao.finditer(texto_debitos):
                    codigo         = m.group(1)
                    nome_parcial   = m.group(2).strip()
                    parcela        = m.group(3)
                    vencimento_str = m.group(4)
                    valor_original = m.group(5)
                    valor_atual    = m.group(6)
                    extra          = m.group(7) or ""
                    # limpa "MATERIAL", "DIDÁTICO" etc.
                    partes = (nome_parcial + " " + extra).split()
                    partes = [p for p in partes if p.upper() not in ("MATERIAL","DIDÁTICO")]
                    nome_completo = " ".join(partes)

                    dados.append({
                        "Código do Aluno": codigo,
                        "Nome do Aluno": nome_completo,
                        "Parcela": parcela,
                        "Vencimento": vencimento_str,
                        # a coluna "Dias de Atraso" vai receber fórmula depois
                        "Dias de Atraso": None,
                        "Valor Original": valor_original,
                        "Valor Atualizado": valor_atual,
                        "Responsável Nome": responsavel["nome"],
                        "Responsável CPF": responsavel["cpf"],
                        "Responsável Email": responsavel["email"],
                    })
                continue

            i += 1

# 1) salva com pandas
df = pd.DataFrame(dados)
arquivo = "inadimplentes.xlsx"
df.to_excel(arquivo, index=False)

# 2) reabre com openpyxl e injeta fórmulas
wb = load_workbook(arquivo)
ws = wb.active

# identifica as colunas
col_venc = col_dias = None
for col in range(1, ws.max_column + 1):
    header = ws.cell(1, col).value
    if header == "Vencimento":
        col_venc = get_column_letter(col)
    elif header == "Dias de Atraso":
        col_dias = get_column_letter(col)

# insere fórmula em cada linha
for row in range(2, ws.max_row + 1):
    # fórmula: =MAX(TODAY()-D2,0)
    cell_dias = f"{col_dias}{row}"
    cell_venc = f"{col_venc}{row}"
    ws[cell_dias] = f"=MAX(TODAY()-{cell_venc},0)"

# salva de volta
wb.save(arquivo)

print(f"✅ '{arquivo}' gerado com sucesso!")
